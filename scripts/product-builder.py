#!/usr/bin/env python3
"""
Cipher Studio — Product File Builder
=====================================
Generates branded .xlsx files from product specs stored in Supabase,
uploads them to Supabase Storage, and returns a public download URL.

USAGE
-----
CLI (run once):
    python product-builder.py --product-id <uuid>

HTTP server (for n8n webhook calls):
    python product-builder.py --serve --port 8000

Deploy on Railway as a separate service alongside n8n.

REQUIREMENTS
------------
pip install -r requirements.txt

ENVIRONMENT VARIABLES
---------------------
SUPABASE_URL
SUPABASE_SECRET_KEY
"""

import os
import sys
import json
import logging
import argparse
from io import BytesIO
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from supabase import create_client, Client

# FastAPI + uvicorn (only needed for --serve mode)
try:
    from fastapi import FastAPI, HTTPException
    from pydantic import BaseModel
    import uvicorn
    FASTAPI_AVAILABLE = True
except ImportError:
    FASTAPI_AVAILABLE = False

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("product-builder")

# ---------------------------------------------------------------------------
# Brand constants
# ---------------------------------------------------------------------------
LIME   = "C8F53A"   # Cipher Lime
BLACK  = "0A0A0A"   # Void Black
WHITE  = "FFFFFF"
GRAY1  = "111111"   # near-black card bg
GRAY2  = "1A1A1A"   # slightly lighter
GRAY3  = "2A2A2A"   # border grey
LGRAY  = "AAAAAA"   # muted text
STRIPE = "F7FDE0"   # lime-tinted alternating row

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _font(color: str = WHITE, bold: bool = False, size: int = 10, italic: bool = False) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def _align(h: str = "left", v: str = "center", wrap: bool = False, indent: int = 0) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def _thin_border() -> Border:
    thin = Side(style="thin", color=GRAY3)
    return Border(bottom=thin)

def _set_header(cell, value: str, *, fg: str = BLACK, bg: str = LIME,
                bold: bool = True, size: int = 10, h: str = "left", indent: int = 1):
    cell.value = value
    cell.font = _font(fg, bold, size)
    cell.fill = _fill(bg)
    cell.alignment = _align(h, indent=indent)

# ---------------------------------------------------------------------------
# Tab builders
# ---------------------------------------------------------------------------
def _build_instructions_tab(wb: openpyxl.Workbook, product_name: str,
                             spec: dict, all_tabs: list):
    ws = wb.create_sheet("Instructions", 0)
    ws.sheet_properties.tabColor = LIME

    # Banner
    ws.merge_cells("A1:G1")
    _set_header(ws["A1"], f"CIPHER STUDIO  —  {product_name.upper()}",
                fg=LIME, bg=BLACK, size=16, indent=1)
    ws.row_dimensions[1].height = 42

    ws.merge_cells("A2:G2")
    ws["A2"].value = "Template Guide & Setup Instructions"
    ws["A2"].font = _font(LGRAY, size=10)
    ws["A2"].fill = _fill(GRAY1)
    ws["A2"].alignment = _align(indent=1)
    ws.row_dimensions[2].height = 22

    row = 4

    # How to use
    ws.merge_cells(f"A{row}:G{row}")
    _set_header(ws[f"A{row}"], "HOW TO USE THIS TEMPLATE")
    ws.row_dimensions[row].height = 22
    row += 1

    steps = spec.get("sections") or [
        "1. Navigate between tabs using the sheet tabs at the bottom.",
        "2. Enter your data in cells with a white or light background.",
        "3. Do NOT edit cells with the lime (#C8F53A) header colour — these contain formulas.",
        "4. The Dashboard tab updates automatically as you enter data.",
        "5. Save a backup copy before making structural changes.",
    ]
    for i, step in enumerate(steps):
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"].value = step
        ws[f"A{row}"].font = _font("222222", size=10)
        ws[f"A{row}"].fill = _fill(STRIPE if i % 2 == 0 else WHITE)
        ws[f"A{row}"].alignment = _align(indent=2, wrap=True)
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1

    # Tab index
    ws.merge_cells(f"A{row}:G{row}")
    _set_header(ws[f"A{row}"], "TABS IN THIS FILE")
    ws.row_dimensions[row].height = 22
    row += 1

    for tab in all_tabs:
        ws[f"A{row}"].value = tab.get("name", "")
        ws[f"A{row}"].font = _font(LIME, bold=True, size=10)
        ws[f"A{row}"].fill = _fill(BLACK)
        ws[f"A{row}"].alignment = _align(indent=1)

        ws.merge_cells(f"B{row}:G{row}")
        ws[f"B{row}"].value = tab.get("purpose", "")
        ws[f"B{row}"].font = _font("CCCCCC", size=10)
        ws[f"B{row}"].fill = _fill(BLACK)
        ws[f"B{row}"].alignment = _align()
        ws.row_dimensions[row].height = 20
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].value = "Built by Cipher Studio  ·  cipherstudio.co"
    ws[f"A{row}"].font = _font(LGRAY, italic=True, size=9)
    ws[f"A{row}"].alignment = _align(h="center")

    ws.column_dimensions["A"].width = 28
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 16


def _build_dashboard_tab(wb: openpyxl.Workbook, product_name: str, spec: dict):
    ws = wb.create_sheet("Dashboard", 1)
    ws.sheet_properties.tabColor = LIME

    ws.merge_cells("A1:H1")
    _set_header(ws["A1"], f"{product_name}  —  Dashboard",
                fg=LIME, bg=BLACK, size=14, indent=1)
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:H2")
    ws["A2"].value = f"Last updated: =NOW()"
    ws["A2"].font = _font(LGRAY, size=9)
    ws["A2"].fill = _fill(GRAY1)
    ws["A2"].alignment = _align(h="right", indent=1)
    ws.row_dimensions[2].height = 18

    # KPI cards (2 per row, 4 total)
    kpis = (spec.get("kpis") or [])[:8]
    if not kpis:
        kpis = ["Total Records", "This Month", "Running Total", "Average"]

    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    _set_header(ws[f"A{row}"], "KEY METRICS")
    ws.row_dimensions[row].height = 22
    row += 1

    kpi_cols = [1, 3, 5, 7]
    for i, kpi in enumerate(kpis[:4]):
        name = kpi.split(":")[0].strip() if ":" in kpi else kpi
        col = kpi_cols[i % 4]

        label = ws.cell(row=row, column=col)
        label.value = name.upper()
        label.font = _font(LGRAY, size=8, bold=True)
        label.fill = _fill(GRAY2)
        label.alignment = _align(h="center")

        value = ws.cell(row=row + 1, column=col)
        value.value = "—"
        value.font = Font(name="Calibri", bold=True, size=22, color=LIME)
        value.fill = _fill(BLACK)
        value.alignment = _align(h="center")
        ws.row_dimensions[row + 1].height = 44

        if (i + 1) % 4 == 0:
            row += 3

    for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 15


def _build_data_tab(wb: openpyxl.Workbook, tab: dict):
    name = (tab.get("name") or "Data")[:31]  # Excel max 31 chars
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = GRAY3

    columns = tab.get("columns") or []
    formulas = tab.get("formulas") or []

    # Column headers
    for col_idx, col_def in enumerate(columns, 1):
        col_name = col_def.split(":")[0].strip() if ":" in col_def else col_def
        cell = ws.cell(row=1, column=col_idx)
        _set_header(cell, col_name, h="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    ws.row_dimensions[1].height = 22

    # Sample data rows with alternating fill
    for data_row in range(2, 52):
        stripe = _fill(STRIPE) if data_row % 2 == 0 else _fill(WHITE)
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = stripe
            cell.font = _font("222222", size=10)
            cell.alignment = _align(v="center")
            cell.border = _thin_border()
        ws.row_dimensions[data_row].height = 18

    # Formula helper column (if formulas defined)
    if formulas:
        formula_col = len(columns) + 1
        header_cell = ws.cell(row=1, column=formula_col)
        _set_header(header_cell, "Calculated", fg=WHITE, bg=GRAY3, h="center")
        ws.column_dimensions[get_column_letter(formula_col)].width = 20

        for data_row in range(2, 52):
            cell = ws.cell(row=data_row, column=formula_col)
            cell.fill = _fill(GRAY2)
            cell.font = _font(LIME, size=10)
            cell.alignment = _align(h="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------
def build_spreadsheet(product_name: str, spec: dict) -> BytesIO:
    """Return an in-memory .xlsx file for the given spreadsheet spec."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    tabs = spec.get("tabs") or []
    dashboard_spec = spec.get("dashboard") or {}
    instructions_spec = spec.get("instructions_tab") or {}

    _build_instructions_tab(wb, product_name, instructions_spec, tabs)
    _build_dashboard_tab(wb, product_name, dashboard_spec)
    for tab in tabs:
        _build_data_tab(wb, tab)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_and_upload(product_id: str) -> dict:
    """
    Full pipeline:
      1. Fetch product + spec from Supabase
      2. Build .xlsx
      3. Upload to Supabase Storage bucket 'products'
      4. Update products row with file_url + status
      5. Return result dict
    """
    url = os.environ["SUPABASE_URL"]
    key = os.environ["SUPABASE_SECRET_KEY"]
    supabase: Client = create_client(url, key)

    # 1. Fetch product
    log.info("Fetching product %s", product_id)
    res = supabase.table("products").select("*").eq("id", product_id).single().execute()
    product = res.data
    if not product:
        raise ValueError(f"Product {product_id} not found in Supabase")

    product_type = product.get("type", "")
    product_name = product.get("name", "Cipher Studio Template")
    spec = product.get("spec") or {}

    if product_type != "spreadsheet":
        raise NotImplementedError(
            f"Type '{product_type}' not yet supported. "
            "Only 'spreadsheet' is implemented in this version."
        )

    # 2. Build file
    log.info("Building .xlsx for: %s", product_name)
    file_buf = build_spreadsheet(product_name, spec)

    # 3. Upload to Supabase Storage
    safe_name = "".join(c if c.isalnum() or c == "-" else "-" for c in product_name.lower())[:50]
    storage_path = f"{product_id}/{safe_name}.xlsx"
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    log.info("Uploading to storage: products/%s", storage_path)
    supabase.storage.from_("products").upload(
        path=storage_path,
        file=file_buf.read(),
        file_options={"content-type": content_type, "upsert": "true"},
    )

    # 4. Get public URL
    public_url = supabase.storage.from_("products").get_public_url(storage_path)
    log.info("Public URL: %s", public_url)

    # 5. Update product record
    supabase.table("products").update({
        "file_url": public_url,
        "status": "ready_for_review",
    }).eq("id", product_id).execute()
    log.info("Product record updated — status: ready_for_review")

    return {
        "success": True,
        "product_id": product_id,
        "product_name": product_name,
        "file_url": public_url,
        "storage_path": storage_path,
    }


# ---------------------------------------------------------------------------
# FastAPI server (Railway deployment / n8n webhook target)
# ---------------------------------------------------------------------------
if FASTAPI_AVAILABLE:
    app = FastAPI(title="Cipher Studio Product Builder", version="1.0.0")

    class BuildRequest(BaseModel):
        product_id: str

    @app.post("/build")
    def build_endpoint(req: BuildRequest):
        try:
            result = build_and_upload(req.product_id)
            return result
        except NotImplementedError as e:
            raise HTTPException(status_code=422, detail=str(e))
        except ValueError as e:
            raise HTTPException(status_code=404, detail=str(e))
        except Exception as e:
            log.exception("Unhandled error building product %s", req.product_id)
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/health")
    def health():
        return {"status": "ok", "service": "cipher-studio-product-builder"}


# ---------------------------------------------------------------------------
# CLI entrypoint
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    # Load .env if present (for local dev)
    try:
        from dotenv import load_dotenv
        load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "..", ".env"))
    except ImportError:
        pass

    parser = argparse.ArgumentParser(
        description="Cipher Studio Product Builder",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--product-id", metavar="UUID",
                        help="Build a specific product by its Supabase UUID")
    parser.add_argument("--serve", action="store_true",
                        help="Run as an HTTP API server (requires fastapi + uvicorn)")
    parser.add_argument("--port", type=int, default=8000,
                        help="Port for HTTP server (default: 8000)")
    args = parser.parse_args()

    if args.serve:
        if not FASTAPI_AVAILABLE:
            sys.exit("Error: fastapi and uvicorn are required for --serve mode. "
                     "Run: pip install fastapi uvicorn")
        log.info("Starting Cipher Studio Product Builder on port %d", args.port)
        uvicorn.run(app, host="0.0.0.0", port=args.port)

    elif args.product_id:
        result = build_and_upload(args.product_id)
        print(json.dumps(result, indent=2))

    else:
        parser.print_help()
