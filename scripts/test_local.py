#!/usr/bin/env python3
"""
Local test — Freelancer Monthly Budget Tracker
Generates test-outputs/freelancer-budget-tracker.xlsx with real formulas.
Run: python3 scripts/test_local.py
"""

import os
import sys
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

# ---------------------------------------------------------------------------
# Brand
# ---------------------------------------------------------------------------
LIME   = "C8F53A"
BLACK  = "0A0A0A"
WHITE  = "FFFFFF"
LGRAY  = "F0F0F0"   # alternating row (light gray)
PANEL  = "F8F8F8"   # card / section background
BORDER = "DDDDDD"
MUTED  = "888888"
TEXT   = "222222"

def _fill(c): return PatternFill(start_color=c, end_color=c, fill_type="solid")
def _font(c=TEXT, bold=False, sz=10, italic=False):
    return Font(name="Calibri", bold=bold, color=c, size=sz, italic=italic)
def _align(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)
def _border(sides="bottom"):
    thin = Side(style="thin", color=BORDER)
    return Border(
        bottom=thin if "bottom" in sides else None,
        right=thin  if "right"  in sides else None,
    )

def hdr(ws, row, col, val, bg=LIME, fg=BLACK, bold=True, sz=10, h="center", span=0):
    """Write a styled header cell, merging `span` columns if given."""
    cell = ws.cell(row=row, column=col)
    cell.value = val
    cell.font = _font(fg, bold, sz)
    cell.fill = _fill(bg)
    cell.alignment = _align(h, indent=0 if h != "left" else 1)
    if span > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=col + span - 1
        )
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, 22)
    return cell

def dat(ws, row, col, val, stripe=False, fmt=None, h="left", bold=False):
    """Write a regular data cell with alternating stripe fill."""
    cell = ws.cell(row=row, column=col)
    cell.value = val
    cell.font = _font(TEXT, bold=bold)
    cell.fill = _fill(LGRAY if stripe else WHITE)
    cell.alignment = _align(h)
    cell.border = _border("bottom")
    if fmt:
        cell.number_format = fmt
    return cell

def kpi(ws, row, label_col, val_col, label, value, stripe=False):
    """Write a KPI label + value pair."""
    lc = ws.cell(row=row, column=label_col)
    lc.value = label
    lc.font = _font(TEXT, bold=True)
    lc.fill = _fill(LGRAY if stripe else WHITE)
    lc.alignment = _align("left", indent=1)
    lc.border = _border("bottom")

    vc = ws.cell(row=row, column=val_col)
    vc.value = value
    vc.font = _font(BLACK, bold=True, sz=11)
    vc.fill = _fill(LGRAY if stripe else WHITE)
    vc.alignment = _align("right")
    vc.border = _border("bottom right")
    vc.number_format = '"$"#,##0.00'
    ws.row_dimensions[row].height = 20
    return vc


# ---------------------------------------------------------------------------
# Tab 1: Dashboard
# ---------------------------------------------------------------------------
def create_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = LIME

    # ── Banner ──────────────────────────────────────────────────────────────
    hdr(ws, 1, 1, "FREELANCER MONTHLY BUDGET TRACKER",
        bg=BLACK, fg=LIME, sz=16, h="left", span=6)
    ws.row_dimensions[1].height = 44

    hdr(ws, 2, 1, f"Cipher Studio  ·  cipherstudio.co  ·  {date.today().strftime('%B %Y')}",
        bg=BLACK, fg=MUTED, sz=9, h="left", span=6, bold=False)
    ws.row_dimensions[2].height = 18

    # ── Monthly Summary ──────────────────────────────────────────────────────
    hdr(ws, 4, 1, "MONTHLY SUMMARY", bg=LIME, fg=BLACK, span=6)

    # KPI rows — formulas reference Income Tracker and Expense Tracker
    kpi(ws, 5, 1, 2, "Total Income (Paid)",
        "=SUMIF('Income Tracker'!F2:F1001,\"Paid\",'Income Tracker'!D2:D1001)")
    kpi(ws, 6, 1, 2, "Total Expenses",
        "=SUM('Expense Tracker'!C2:C1001)", stripe=True)
    kpi(ws, 7, 1, 2, "Net Profit",
        "=B5-B6")
    ws.cell(row=7, column=2).font = _font(BLACK, bold=True, sz=12)
    ws.cell(row=7, column=2).fill = _fill(LIME)

    kpi(ws, 8, 1, 2, "Tax Set Aside (30%)",
        "=MAX(0,B7*0.3)", stripe=True)
    kpi(ws, 9, 1, 2, "Spendable Take-Home",
        "=B7-B8")

    # ── Pending & Overdue ───────────────────────────────────────────────────
    ws.row_dimensions[11].height = 6
    hdr(ws, 12, 1, "INVOICE STATUS", bg=LIME, fg=BLACK, span=6)

    kpi(ws, 13, 1, 2, "Pending Invoices (value)",
        "=SUMIF('Income Tracker'!F2:F1001,\"Pending\",'Income Tracker'!D2:D1001)")
    kpi(ws, 14, 1, 2, "Overdue Invoices (value)",
        "=SUMIF('Income Tracker'!F2:F1001,\"Overdue\",'Income Tracker'!D2:D1001)",
        stripe=True)
    kpi(ws, 15, 1, 2, "Tax-Deductible Expenses",
        "=SUMIF('Expense Tracker'!E2:E1001,\"Yes\",'Expense Tracker'!C2:C1001)")

    # ── Tax Quick-View ──────────────────────────────────────────────────────
    ws.row_dimensions[17].height = 6
    hdr(ws, 18, 1, "QUARTERLY TAX ESTIMATES", bg=LIME, fg=BLACK, span=6)

    for i, qtr in enumerate(["Q1 (Jan–Mar)", "Q2 (Apr–Jun)", "Q3 (Jul–Sep)", "Q4 (Oct–Dec)"]):
        r = 19 + i
        kpi(ws, r, 1, 2, f"{qtr} Estimated Tax",
            f"='Tax Estimator'!B9", stripe=(i % 2 == 0))

    # ── Income vs Expenses chart ─────────────────────────────────────────────
    # Reference table placed in columns H-I, rows 4-7 (clear of merged cells)
    ws.cell(4, 8).value = "Category"
    ws.cell(4, 9).value = "Amount"
    ws.cell(4, 8).font = _font(BLACK, bold=True)
    ws.cell(4, 9).font = _font(BLACK, bold=True)
    ws.cell(4, 8).fill = _fill(LIME)
    ws.cell(4, 9).fill = _fill(LIME)
    ws.cell(5, 8).value = "Income"
    ws.cell(5, 9).value = "=B5"
    ws.cell(6, 8).value = "Expenses"
    ws.cell(6, 9).value = "=B6"
    ws.cell(7, 8).value = "Net Profit"
    ws.cell(7, 9).value = "=B7"
    for r in [5, 6, 7]:
        ws.cell(r, 8).font = _font(TEXT)
        ws.cell(r, 9).font = _font(TEXT)
        ws.cell(r, 9).number_format = '"$"#,##0.00'

    chart = BarChart()
    chart.type = "col"
    chart.title = "Income vs Expenses"
    chart.style = 2
    chart.grouping = "clustered"
    chart.y_axis.title = "Amount ($)"
    chart.x_axis.title = None
    chart.height = 10
    chart.width = 14

    data_ref = Reference(ws, min_col=9, min_row=4, max_row=7)
    cats_ref = Reference(ws, min_col=8, min_row=5, max_row=7)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = LIME
    chart.series[0].graphicalProperties.line.solidFill = BLACK
    ws.add_chart(chart, "D22")

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 2
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 18


# ---------------------------------------------------------------------------
# Tab 2: Income Tracker
# ---------------------------------------------------------------------------
def create_income_tracker(wb):
    ws = wb.create_sheet("Income Tracker")
    ws.sheet_properties.tabColor = "333333"

    COLS = ["Client", "Project", "Invoice #", "Amount ($)", "Date Paid", "Status"]
    WIDTHS = [24, 28, 14, 14, 14, 12]

    # Banner
    hdr(ws, 1, 1, "INCOME TRACKER", bg=BLACK, fg=LIME, sz=13, h="left",
        span=len(COLS), bold=True)
    ws.row_dimensions[1].height = 32

    # Column headers
    for c, (label, w) in enumerate(zip(COLS, WIDTHS), 1):
        hdr(ws, 2, c, label, bg=LIME, fg=BLACK, sz=10)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[2].height = 22

    # Totals row (row 3, above data — pinned summary)
    ws.merge_cells("A3:C3")
    ws["A3"].value = "TOTALS"
    ws["A3"].font = _font(WHITE, bold=True, sz=10)
    ws["A3"].fill = _fill(BLACK)
    ws["A3"].alignment = _align("right")
    ws["D3"].value = '=SUMIF(F4:F1001,"Paid",D4:D1001)'
    ws["D3"].font = _font(LIME, bold=True, sz=11)
    ws["D3"].fill = _fill(BLACK)
    ws["D3"].number_format = '"$"#,##0.00'
    ws["D3"].alignment = _align("center")
    ws["E3"].value = "Paid total"
    ws["E3"].font = _font(MUTED, sz=9, italic=True)
    ws["E3"].fill = _fill(BLACK)
    ws["F3"].fill = _fill(BLACK)
    ws.row_dimensions[3].height = 22

    # Sample data
    sample = [
        ("Acme Corp",      "Website Redesign",  "INV-001", 3500.00, "2024-01-15", "Paid"),
        ("Jane's Bakery",  "Logo Design",        "INV-002", 850.00,  "2024-01-28", "Paid"),
        ("Startup Inc.",   "Brand Strategy",     "INV-003", 2200.00, "2024-02-05", "Paid"),
        ("Blue Oak Studio","Content Writing",    "INV-004", 650.00,  "2024-02-20", "Pending"),
        ("TechFlow Ltd.",  "UI/UX Audit",        "INV-005", 1800.00, "",           "Overdue"),
    ]
    for i, row_data in enumerate(sample):
        r = i + 4
        stripe = (i % 2 == 1)
        dat(ws, r, 1, row_data[0], stripe)
        dat(ws, r, 2, row_data[1], stripe)
        dat(ws, r, 3, row_data[2], stripe, h="center")
        dat(ws, r, 4, row_data[3], stripe, fmt='"$"#,##0.00', h="right")
        dat(ws, r, 5, row_data[4], stripe, fmt="YYYY-MM-DD", h="center")
        # Status with conditional colour
        sc = ws.cell(row=r, column=6)
        sc.value = row_data[5]
        sc.font = _font(
            BLACK if row_data[5] == "Paid" else
            ("CC6600" if row_data[5] == "Pending" else "CC0000"),
            bold=True
        )
        sc.fill = _fill(
            LIME  if row_data[5] == "Paid" else
            ("FFF3CD" if row_data[5] == "Pending" else "FDECEA")
        )
        sc.alignment = _align("center")
        sc.border = _border("bottom")
        ws.row_dimensions[r].height = 20

    # Blank input rows
    for i in range(len(sample), 50):
        r = i + 4
        stripe = (i % 2 == 1)
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.fill = _fill(LGRAY if stripe else WHITE)
            cell.border = _border("bottom")
            if c == 4:
                cell.number_format = '"$"#,##0.00'
            if c == 5:
                cell.number_format = "YYYY-MM-DD"
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A2:F2"


# ---------------------------------------------------------------------------
# Tab 3: Expense Tracker
# ---------------------------------------------------------------------------
def create_expense_tracker(wb):
    ws = wb.create_sheet("Expense Tracker")
    ws.sheet_properties.tabColor = "333333"

    COLS = ["Category", "Vendor / Description", "Amount ($)", "Date", "Tax Deductible?"]
    WIDTHS = [22, 34, 14, 14, 18]

    hdr(ws, 1, 1, "EXPENSE TRACKER", bg=BLACK, fg=LIME, sz=13, h="left",
        span=len(COLS), bold=True)
    ws.row_dimensions[1].height = 32

    for c, (label, w) in enumerate(zip(COLS, WIDTHS), 1):
        hdr(ws, 2, c, label, bg=LIME, fg=BLACK, sz=10)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[2].height = 22

    # Totals row
    ws.merge_cells("A3:B3")
    ws["A3"].value = "TOTALS"
    ws["A3"].font = _font(WHITE, bold=True)
    ws["A3"].fill = _fill(BLACK)
    ws["A3"].alignment = _align("right")
    ws["C3"].value = "=SUM(C4:C1001)"
    ws["C3"].font = _font(LIME, bold=True, sz=11)
    ws["C3"].fill = _fill(BLACK)
    ws["C3"].number_format = '"$"#,##0.00'
    ws["C3"].alignment = _align("center")
    ws["D3"].value = "Tax deductible →"
    ws["D3"].font = _font(MUTED, sz=9, italic=True)
    ws["D3"].fill = _fill(BLACK)
    ws["E3"].value = '=SUMIF(E4:E1001,"Yes",C4:C1001)'
    ws["E3"].font = _font(LIME, bold=True, sz=11)
    ws["E3"].fill = _fill(BLACK)
    ws["E3"].number_format = '"$"#,##0.00'
    ws["E3"].alignment = _align("center")
    ws.row_dimensions[3].height = 22

    sample = [
        ("Software & Tools", "Adobe Creative Cloud",  54.99,  "2024-01-02", "Yes"),
        ("Software & Tools", "Notion Pro",             16.00,  "2024-01-02", "Yes"),
        ("Home Office",      "Desk ergonomic chair",  349.00,  "2024-01-10", "Yes"),
        ("Marketing",        "LinkedIn Premium",       39.99,  "2024-01-15", "Yes"),
        ("Meals & Client",   "Client lunch - Acme",    62.50,  "2024-01-16", "Yes"),
        ("Education",        "Udemy course bundle",    29.99,  "2024-01-20", "Yes"),
        ("Utilities",        "Internet bill",          89.00,  "2024-02-01", "No"),
        ("Travel",           "Uber to client meeting", 24.75,  "2024-02-08", "Yes"),
    ]
    for i, row_data in enumerate(sample):
        r = i + 4
        stripe = (i % 2 == 1)
        dat(ws, r, 1, row_data[0], stripe)
        dat(ws, r, 2, row_data[1], stripe)
        dat(ws, r, 3, row_data[2], stripe, fmt='"$"#,##0.00', h="right")
        dat(ws, r, 4, row_data[3], stripe, fmt="YYYY-MM-DD", h="center")
        tc = ws.cell(row=r, column=5)
        tc.value = row_data[4]
        tc.font = _font(BLACK if row_data[4] == "Yes" else MUTED, bold=(row_data[4] == "Yes"))
        tc.fill = _fill(LIME if row_data[4] == "Yes" else (LGRAY if stripe else WHITE))
        tc.alignment = _align("center")
        tc.border = _border("bottom")
        ws.row_dimensions[r].height = 20

    for i in range(len(sample), 50):
        r = i + 4
        stripe = (i % 2 == 1)
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.fill = _fill(LGRAY if stripe else WHITE)
            cell.border = _border("bottom")
            if c == 3:
                cell.number_format = '"$"#,##0.00'
            if c == 4:
                cell.number_format = "YYYY-MM-DD"
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A2:E2"


# ---------------------------------------------------------------------------
# Tab 4: Tax Estimator
# ---------------------------------------------------------------------------
def create_tax_estimator(wb):
    ws = wb.create_sheet("Tax Estimator")
    ws.sheet_properties.tabColor = LIME

    hdr(ws, 1, 1, "QUARTERLY TAX ESTIMATOR", bg=BLACK, fg=LIME, sz=13, h="left",
        span=3, bold=True)
    ws.row_dimensions[1].height = 36

    hdr(ws, 2, 1,
        "Auto-calculated from your Income and Expense Tracker — update those tabs first",
        bg=BLACK, fg=MUTED, sz=9, h="left", span=3, bold=False)
    ws.row_dimensions[2].height = 18

    def section(row, label):
        hdr(ws, row, 1, label, bg=LIME, fg=BLACK, span=3)
        ws.row_dimensions[row].height = 22

    def row_pair(row, label, formula_or_val, fmt='"$"#,##0.00', bold=False, stripe=False,
                 note=""):
        lc = ws.cell(row=row, column=1)
        lc.value = label
        lc.font = _font(TEXT, bold=bold)
        lc.fill = _fill(LGRAY if stripe else WHITE)
        lc.alignment = _align("left", indent=1)
        lc.border = _border("bottom")

        vc = ws.cell(row=row, column=2)
        vc.value = formula_or_val
        vc.font = _font(BLACK, bold=bold, sz=11 if bold else 10)
        vc.fill = _fill(LIME if bold else (LGRAY if stripe else WHITE))
        vc.alignment = _align("right")
        vc.border = _border("bottom right")
        if fmt:
            vc.number_format = fmt

        nc = ws.cell(row=row, column=3)
        nc.value = note
        nc.font = _font(MUTED, sz=9, italic=True)
        nc.fill = _fill(LGRAY if stripe else WHITE)
        nc.alignment = _align("left", indent=1)
        nc.border = _border("bottom")
        ws.row_dimensions[row].height = 20

    # ── Income section ────────────────────────────────────────────────────
    section(4, "INCOME & EXPENSES (Year-to-Date)")
    row_pair(5, "Gross Revenue (all invoices sent)",
             "=SUMIF('Income Tracker'!F2:F1001,\"Paid\",'Income Tracker'!D2:D1001)",
             note="Paid invoices only")
    row_pair(6, "Tax-Deductible Business Expenses",
             "=SUMIF('Expense Tracker'!E2:E1001,\"Yes\",'Expense Tracker'!C2:C1001)",
             stripe=True, note="Expenses marked Yes")
    row_pair(7, "Net Taxable Income",
             "=B5-B6", bold=True, note="Revenue minus deductions")

    # ── Tax calculation ────────────────────────────────────────────────────
    ws.row_dimensions[9].height = 8
    section(10, "TAX CALCULATION")

    # Self-employment tax: 15.3% on 92.35% of net (IRS formula)
    row_pair(11, "Self-Employment Tax (15.3%)",
             "=B7*0.9235*0.153", note="92.35% of net × 15.3%")
    # SE tax deduction (50% of SE tax is deductible)
    row_pair(12, "SE Tax Deduction (50% of SE Tax)",
             "=B11*0.5", stripe=True, note="Reduces income tax basis")
    row_pair(13, "Adjusted Gross Income",
             "=B7-B12", note="After SE deduction")
    # Federal income tax — 22% bracket estimate (common for freelancers)
    row_pair(14, "Federal Income Tax (22% est.)",
             "=B13*0.22", stripe=True, note="Adjust rate for your bracket")
    row_pair(15, "Total Annual Tax Estimate",
             "=B11+B14", bold=True)

    # ── Quarterly breakdown ─────────────────────────────────────────────────
    ws.row_dimensions[17].height = 8
    section(18, "QUARTERLY ESTIMATED PAYMENTS (Due to IRS)")
    row_pair(19, "Q1 — Due April 15",    "=B15/4", stripe=False, note="Jan 1 – Mar 31")
    row_pair(20, "Q2 — Due June 15",     "=B15/4", stripe=True,  note="Apr 1 – May 31")
    row_pair(21, "Q3 — Due September 15","=B15/4", stripe=False, note="Jun 1 – Aug 31")
    row_pair(22, "Q4 — Due January 15",  "=B15/4", stripe=True,  note="Sep 1 – Dec 31")

    # ── Reminder ───────────────────────────────────────────────────────────
    ws.row_dimensions[24].height = 8
    hdr(ws, 25, 1,
        "⚠  This is an estimate. Consult a tax professional for your final filing.",
        bg="FFF3CD", fg="856404", sz=9, h="left", span=3, bold=False)
    ws.row_dimensions[25].height = 20

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30


# ---------------------------------------------------------------------------
# Tab 5: Instructions (first sheet)
# ---------------------------------------------------------------------------
def create_instructions(wb):
    ws = wb.create_sheet("Start Here")
    ws.sheet_properties.tabColor = LIME

    hdr(ws, 1, 1, "CIPHER STUDIO", bg=BLACK, fg=LIME, sz=11, h="left", span=2, bold=True)
    hdr(ws, 2, 1, "FREELANCER MONTHLY BUDGET TRACKER", bg=BLACK, fg=WHITE, sz=18,
        h="left", span=2, bold=True)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 46

    hdr(ws, 3, 1, "The complete financial tracking system for freelancers — income, expenses, and taxes in one file.",
        bg=BLACK, fg=MUTED, sz=10, h="left", span=2, bold=False)
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 10

    hdr(ws, 5, 1, "HOW TO USE THIS TEMPLATE", bg=LIME, fg=BLACK, span=2)
    ws.row_dimensions[5].height = 22

    steps = [
        ("1.  Log income",     "Go to Income Tracker → add each invoice when sent. Mark Status as Paid when payment arrives."),
        ("2.  Log expenses",   "Go to Expense Tracker → add every business expense. Mark Tax Deductible? as Yes when applicable."),
        ("3.  Check dashboard","Dashboard auto-calculates income, expenses, net profit, and tax set-aside. No manual entry needed."),
        ("4.  Review taxes",   "Tax Estimator pulls your live numbers and calculates quarterly IRS payments automatically."),
        ("5.  Save monthly",   "Duplicate this file each month or use one file per year. Keep backups before editing structure."),
    ]
    for i, (title, body) in enumerate(steps):
        r = 6 + i
        stripe = (i % 2 == 1)
        ws.cell(r, 1).value = title
        ws.cell(r, 1).font = _font(BLACK, bold=True)
        ws.cell(r, 1).fill = _fill(LGRAY if stripe else WHITE)
        ws.cell(r, 1).alignment = _align("left", indent=1)
        ws.cell(r, 1).border = _border("bottom")
        ws.cell(r, 2).value = body
        ws.cell(r, 2).font = _font(TEXT)
        ws.cell(r, 2).fill = _fill(LGRAY if stripe else WHITE)
        ws.cell(r, 2).alignment = _align("left", wrap=True)
        ws.cell(r, 2).border = _border("bottom")
        ws.row_dimensions[r].height = 24

    ws.row_dimensions[12].height = 10
    hdr(ws, 13, 1, "TABS IN THIS FILE", bg=LIME, fg=BLACK, span=2)
    ws.row_dimensions[13].height = 22

    tabs_info = [
        ("Dashboard",        "Live summary: income, expenses, net profit, tax status, and chart."),
        ("Income Tracker",   "Log every client invoice — amount, date, and payment status."),
        ("Expense Tracker",  "Track every business expense and flag tax-deductible items."),
        ("Tax Estimator",    "Auto-calculates federal tax and quarterly IRS payment amounts."),
    ]
    for i, (name, desc) in enumerate(tabs_info):
        r = 14 + i
        stripe = (i % 2 == 1)
        ws.cell(r, 1).value = name
        ws.cell(r, 1).font = _font(BLACK, bold=True, sz=10)
        ws.cell(r, 1).fill = _fill(LIME if not stripe else LGRAY)
        ws.cell(r, 1).alignment = _align("left", indent=1)
        ws.cell(r, 1).border = _border("bottom")
        ws.cell(r, 2).value = desc
        ws.cell(r, 2).font = _font(TEXT)
        ws.cell(r, 2).fill = _fill(LGRAY if stripe else WHITE)
        ws.cell(r, 2).alignment = _align("left", indent=1)
        ws.cell(r, 2).border = _border("bottom")
        ws.row_dimensions[r].height = 20

    ws.row_dimensions[19].height = 14
    ws.cell(20, 1).value = "Built by Cipher Studio  ·  cipherstudio.co  ·  $9 template"
    ws.cell(20, 1).font = _font(MUTED, italic=True, sz=9)
    ws.merge_cells("A20:B20")
    ws.cell(20, 1).alignment = _align("center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 58


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    out_dir = os.path.join(os.path.dirname(__file__), "..", "test-outputs")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "freelancer-budget-tracker.xlsx")

    print("Building Freelancer Monthly Budget Tracker...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_instructions(wb)   # sheet 0 — Start Here
    create_dashboard(wb)      # sheet 1 — Dashboard
    create_income_tracker(wb) # sheet 2 — Income Tracker
    create_expense_tracker(wb)# sheet 3 — Expense Tracker
    create_tax_estimator(wb)  # sheet 4 — Tax Estimator

    wb.save(out_path)
    size_kb = os.path.getsize(out_path) // 1024
    print(f"✓  Saved: {out_path}  ({size_kb} KB)")
    print(f"   Tabs: {[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main()
